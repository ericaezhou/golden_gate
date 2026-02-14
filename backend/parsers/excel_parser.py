"""
Excel parser for .xlsx and .xls files.
Extracts formulas, metadata, comments, and data as markdown tables.
"""

import os
import re
from typing import List, Dict, Any
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from backend.parsers import register_parser, ParseResult


@register_parser("xlsx", "xls")
def parse_excel(path: str) -> ParseResult:
    """
    Parse Excel file and extract formulas, data, metadata, and references.

    Args:
        path: Path to the Excel file

    Returns:
        ParseResult with content, metadata, references, and warnings
    """
    warnings = []
    references = set()
    sheets_metadata = []
    content_parts = []

    try:
        # Load with data_only=False to get formulas
        wb = load_workbook(path, data_only=False, keep_vba=False, keep_links=True)
    except Exception as e:
        return ParseResult(
            file_id=os.path.basename(path),
            file_path=os.path.abspath(path),
            file_type="xlsx",
            content="",
            metadata={},
            references=[],
            warnings=[f"Failed to load workbook: {str(e)}"]
        )

    # Extract named ranges
    named_ranges = []
    try:
        for name in wb.defined_names.values():
            named_ranges.append(f"{name.name}: {name.attr_text}")
    except Exception as e:
        warnings.append(f"Error extracting named ranges: {str(e)}")

    # Add named ranges to content if they exist
    if named_ranges:
        content_parts.append("# Named Ranges\n")
        for nr in named_ranges:
            content_parts.append(f"- {nr}\n")
        content_parts.append("\n")

    # Get hidden sheets
    hidden_sheets = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if sheet.sheet_state == 'hidden' or sheet.sheet_state == 'veryHidden':
            hidden_sheets.append(sheet_name)

    # Process all sheets (including hidden ones)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        is_hidden = sheet_name in hidden_sheets

        # Process sheet
        sheet_data = _process_sheet(sheet, sheet_name, is_hidden, warnings, references)
        sheets_metadata.append(sheet_data['metadata'])
        content_parts.append(sheet_data['content'])

    # Build final metadata
    metadata = {
        "sheets": sheets_metadata,
        "named_ranges": named_ranges,
        "hidden_sheets": hidden_sheets
    }

    # Combine content
    content = "".join(content_parts)

    # Convert references set to sorted list
    references_list = sorted(list(references))

    return ParseResult(
        file_id=os.path.basename(path),
        file_path=os.path.abspath(path),
        file_type="xlsx",
        content=content,
        metadata=metadata,
        references=references_list,
        warnings=warnings
    )


def _process_sheet(sheet: Worksheet, sheet_name: str, is_hidden: bool,
                   warnings: List[str], references: set) -> Dict[str, Any]:
    """
    Process a single worksheet and extract data, formulas, and comments.

    Args:
        sheet: The worksheet to process
        sheet_name: Name of the sheet
        is_hidden: Whether the sheet is hidden
        warnings: List to append warnings to
        references: Set to add external references to

    Returns:
        Dict with 'metadata' and 'content' keys
    """
    content_parts = []
    comments_list = []
    formula_count = 0
    has_formulas = False

    # Sheet header
    hidden_marker = " (hidden)" if is_hidden else ""
    content_parts.append(f"## Sheet: {sheet_name}{hidden_marker}\n\n")

    # Get dimensions
    dimensions = sheet.dimensions if sheet.dimensions else "A1:A1"

    # Get actual data range
    rows_with_data = []
    max_col = 0

    try:
        # Iterate through all rows to find non-empty ones
        for row_idx, row in enumerate(sheet.iter_rows(), start=1):
            row_data = []
            row_has_data = False

            for col_idx, cell in enumerate(row, start=1):
                value = cell.value

                # Check if cell has a comment
                if cell.comment:
                    comment_text = cell.comment.text if hasattr(cell.comment, 'text') else str(cell.comment)
                    comments_list.append(f"{cell.coordinate}: {comment_text}")

                # Check if cell has a formula
                if hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    has_formulas = True
                    formula_count += 1

                    # Extract external references from formula
                    external_refs = re.findall(r'\[([^\]]+\.xlsx?)\]', formula)
                    references.update(external_refs)

                    row_data.append(formula)
                    row_has_data = True
                elif value is not None and value != "":
                    row_data.append(str(value))
                    row_has_data = True
                else:
                    row_data.append("")

                if row_has_data:
                    max_col = max(max_col, col_idx)

            if row_has_data:
                rows_with_data.append((row_idx, row_data))

            # Limit to 100 rows
            if len(rows_with_data) >= 100:
                warnings.append(f"Sheet '{sheet_name}' truncated to 100 rows")
                break

    except Exception as e:
        warnings.append(f"Error processing sheet '{sheet_name}': {str(e)}")
        rows_with_data = []

    row_count = len(rows_with_data)

    # Build metadata for this sheet
    sheet_metadata = {
        "name": sheet_name,
        "dimensions": dimensions,
        "row_count": row_count,
        "has_formulas": has_formulas,
        "formula_cells": formula_count,
        "comments": comments_list
    }

    # Render data as markdown table
    if row_count == 0:
        content_parts.append("*No data in this sheet*\n\n")
        warnings.append(f"Sheet '{sheet_name}' has no data")
    else:
        # Build table header
        header_cols = [get_column_letter(i) for i in range(1, max_col + 1)]
        content_parts.append("| " + " | ".join(header_cols) + " |\n")
        content_parts.append("|" + "|".join(["---"] * max_col) + "|\n")

        # Build table rows
        for row_idx, row_data in rows_with_data:
            # Pad row to max_col length
            while len(row_data) < max_col:
                row_data.append("")

            # Truncate to max_col
            row_data = row_data[:max_col]

            # Escape pipe characters and clean cell values
            escaped_data = []
            for cell_val in row_data:
                # Escape pipes and clean newlines
                cleaned = str(cell_val).replace("|", "\\|").replace("\n", " ").replace("\r", "")
                escaped_data.append(cleaned)

            content_parts.append("| " + " | ".join(escaped_data) + " |\n")

        content_parts.append("\n")

    # Add comments as footnotes
    if comments_list:
        content_parts.append("### Comments\n\n")
        for comment in comments_list:
            content_parts.append(f"- {comment}\n")
        content_parts.append("\n")

    return {
        "metadata": sheet_metadata,
        "content": "".join(content_parts)
    }
